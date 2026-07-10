"""Department-level investigation report export generators."""

from __future__ import annotations

import csv
import io
from datetime import UTC

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from sentinel_anpr.application.dto.investigation_reports_dto import (
    InvestigationReportExportBundleDto,
    InvestigationReportExportResult,
)
from sentinel_anpr.application.ports.outbound.tabular_export_port import TabularExportPort

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
except ImportError:  # pragma: no cover - covered by runtime dependency install
    Workbook = None
    Font = None
    PatternFill = None

_NAVY = colors.HexColor("#0B1F3A")
_BORDER = colors.HexColor("#CBD5E1")
_PANEL = colors.HexColor("#F8FAFC")


class DepartmentInvestigationExportGenerator(TabularExportPort):
    """Render department-level PDF, CSV, and Excel investigation exports."""

    def export_pdf(
        self, bundle: InvestigationReportExportBundleDto
    ) -> InvestigationReportExportResult:
        buffer = io.BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=0.45 * inch,
            rightMargin=0.45 * inch,
            topMargin=0.45 * inch,
            bottomMargin=0.45 * inch,
            title="Department Investigation Report",
            author="SentinelANPR AI",
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            fontSize=18,
            textColor=_NAVY,
            spaceAfter=6,
        )
        body_style = ParagraphStyle(
            "Body",
            parent=styles["Normal"],
            fontSize=9,
            textColor=colors.HexColor("#334155"),
            leading=12,
        )
        story: list[object] = [
            Paragraph("Department Investigation Report", title_style),
            Paragraph(
                "Generated from existing verification history with applied operational filters.",
                body_style,
            ),
            Spacer(1, 10),
            Paragraph(bundle.analytics.investigation_summary, body_style),
            Spacer(1, 8),
            self._summary_table(bundle),
            Spacer(1, 10),
            Paragraph("Filtered Investigations", styles["Heading2"]),
            Spacer(1, 6),
            self._rows_table(bundle),
        ]
        document.build(story)
        return InvestigationReportExportResult(
            filename="department-investigation-report.pdf",
            content_type="application/pdf",
            content=buffer.getvalue(),
        )

    def export_csv(
        self, bundle: InvestigationReportExportBundleDto
    ) -> InvestigationReportExportResult:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Department Investigation Report"])
        writer.writerow(["Generated At", bundle.generated_at.astimezone(UTC).isoformat()])
        writer.writerow(["Summary", bundle.analytics.investigation_summary])
        writer.writerow([])
        writer.writerow(
            [
                "Date",
                "Time",
                "Investigation ID",
                "Registration Number",
                "Owner",
                "Vehicle",
                "Brand",
                "Model",
                "Officer",
                "District",
                "Police Station",
                "Risk Score",
                "Risk Level",
                "Investigation Status",
                "Verification Status",
                "AI Confidence",
                "Report Download",
            ]
        )
        for row in bundle.rows:
            writer.writerow(
                [
                    row.date,
                    row.time,
                    row.investigation_id,
                    row.registration_number,
                    row.owner,
                    row.vehicle,
                    row.brand,
                    row.model,
                    row.officer,
                    row.district,
                    row.police_station,
                    row.risk_score,
                    row.risk_level,
                    row.investigation_status,
                    row.verification_status,
                    row.ai_confidence,
                    row.report_download,
                ]
            )
        return InvestigationReportExportResult(
            filename="department-investigation-report.csv",
            content_type="text/csv; charset=utf-8",
            content=output.getvalue().encode("utf-8"),
        )

    def export_excel(
        self, bundle: InvestigationReportExportBundleDto
    ) -> InvestigationReportExportResult:
        if Workbook is None or Font is None or PatternFill is None:
            raise RuntimeError(
                "openpyxl is not installed. Install with: pip install openpyxl"
            )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Investigations"
        sheet.append(["Department Investigation Report"])
        sheet.append([bundle.analytics.investigation_summary])
        sheet.append([])
        headers = [
            "Date",
            "Time",
            "Investigation ID",
            "Registration Number",
            "Owner",
            "Vehicle",
            "Brand",
            "Model",
            "Officer",
            "District",
            "Police Station",
            "Risk Score",
            "Risk Level",
            "Investigation Status",
            "Verification Status",
            "AI Confidence",
            "Report Download",
        ]
        sheet.append(headers)
        for cell in sheet[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0B1F3A")
        for row in bundle.rows:
            sheet.append(
                [
                    row.date,
                    row.time,
                    row.investigation_id,
                    row.registration_number,
                    row.owner,
                    row.vehicle,
                    row.brand,
                    row.model,
                    row.officer,
                    row.district,
                    row.police_station,
                    row.risk_score,
                    row.risk_level,
                    row.investigation_status,
                    row.verification_status,
                    row.ai_confidence,
                    row.report_download,
                ]
            )
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 32)
        output = io.BytesIO()
        workbook.save(output)
        return InvestigationReportExportResult(
            filename="department-investigation-report.xlsx",
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            content=output.getvalue(),
        )

    @staticmethod
    def _summary_table(bundle: InvestigationReportExportBundleDto) -> Table:
        totals = bundle.analytics.totals
        rows = [
            ["Total Investigations", str(totals.total_investigations)],
            ["Verified Vehicles", str(totals.verified_vehicles)],
            ["Suspicious Vehicles", str(totals.suspicious_vehicles)],
            ["High Risk Vehicles", str(totals.high_risk_vehicles)],
            [
                "Average Risk Score",
                f"{totals.average_risk_score * 100:.1f}%"
                if totals.average_risk_score is not None
                else "-",
            ],
            [
                "Average AI Confidence",
                f"{totals.average_ai_confidence * 100:.1f}%"
                if totals.average_ai_confidence is not None
                else "-",
            ],
            [
                "Average Investigation Time",
                f"{totals.average_investigation_time_minutes:.1f} min"
                if totals.average_investigation_time_minutes is not None
                else "-",
            ],
            ["Top Vehicle Type", totals.top_vehicle_type or "-"],
            ["Top Vehicle Brand", totals.top_vehicle_brand or "-"],
            ["Most Active Officer", totals.most_active_officer or "-"],
            ["Most Active Station", totals.most_active_station or "-"],
        ]
        table = Table(rows, colWidths=[2.2 * inch, 1.6 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), _PANEL),
                    ("BOX", (0, 0), (-1, -1), 0.5, _BORDER),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, _BORDER),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        return table

    @staticmethod
    def _rows_table(bundle: InvestigationReportExportBundleDto) -> Table:
        header = [
            "Date",
            "Time",
            "ID",
            "Registration",
            "Owner",
            "Vehicle",
            "Officer",
            "District",
            "Station",
            "Risk",
            "Status",
            "Verify",
            "AI",
        ]
        body = [
            [
                row.date,
                row.time,
                row.investigation_id,
                row.registration_number,
                row.owner,
                row.vehicle,
                row.officer,
                row.district,
                row.police_station,
                f"{row.risk_score} / {row.risk_level}",
                row.investigation_status,
                row.verification_status,
                row.ai_confidence,
            ]
            for row in bundle.rows[:100]
        ]
        data = [header, *body] if body else [header, ["No data", "", "", "", "", "", "", "", "", "", "", ""]]
        table = Table(
            data,
            colWidths=[0.8, 0.8, 1.2, 1.15, 1.25, 1.3, 1.1, 0.9, 1.0, 1.0, 0.9, 0.8],
            repeatRows=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), _NAVY),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("BOX", (0, 0), (-1, -1), 0.5, _BORDER),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, _BORDER),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        return table
